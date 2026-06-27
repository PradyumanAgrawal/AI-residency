import openpyxl, json

wb = openpyxl.load_workbook('Cohort AI Residency-final.xlsx', data_only=True)
ws = wb.active

def cell(r, c):
    v = ws.cell(row=r, column=c).value
    return ('' if v is None else str(v)).strip()

# Per-row curated keywords + a cleaned sector-group label.
# Keyed by row number (2..20). Source of truth for problem/solution/members/startup is the sheet itself.
meta = {
    2:  ("Space Tech",            ["SAR/InSAR radar", "Satellite compute", "Deformation maps"]),
    3:  ("Enterprise AI",         ["Field sales", "Voice intelligence", "Multilingual", "Privacy-first"]),
    4:  ("Fintech",               ["Loan collections", "Debt recovery", "NBFCs"]),
    5:  ("Consumer Tech",         ["Content engineering", "Low token cost"]),
    6:  ("Physical AI & Robotics",["Robotics data", "Multimodal sensing", "Embodied AI"]),
    7:  ("Healthtech",            ["Genomics", "Risk prediction", "India-specific"]),
    8:  ("Consumer Tech",         ["Creator economy", "Content workflow", "Brand matching"]),
    9:  ("Aerospace",             ["Industrial sensor data", "Physics-aware AI", "Failure modes"]),
    10: ("Consumer Tech",         ["Language learning", "Engagement", "Mobile app"]),
    11: ("Geospatial Tech",       ["Railway safety", "Satellite InSAR", "Ground deformation"]),
    12: ("Geospatial Tech",       ["Earth observation", "Developer API", "Geospatial infra"]),
    13: ("Enterprise AI",         ["Fleet safety", "Driver fatigue", "Edge AI"]),
    14: ("Consumer Tech",         ["D2C fashion", "Conversational commerce", "Ad conversion"]),
    15: ("Enterprise AI",         ["AI services", "D2C & SME", "Student teams"]),
    16: ("Enterprise AI",         ["Power grid", "Theft & loss", "Solar curtailment"]),
    17: ("Advanced Manufacturing",["Design feasibility", "Contract manufacturing", "Quoting"]),
    18: ("Fintech",               ["AML / RegTech", "Alert triage", "False positives"]),
    19: ("Consumer Tech",         ["English learning", "Short-form video"]),
    20: ("Regulatory Tech",       ["DPDP Act", "Compliance", "Automation"]),
}

# Display order of sector groups.
# Business-relevant sectors first (for the MBA audience), deep-tech / space last.
group_order = [
    "Enterprise AI", "Consumer Tech", "Regulatory Tech", "Fintech", "Blockchain",
    "Healthtech", "Space Tech", "Aerospace", "Geospatial Tech",
    "Physical AI & Robotics", "Advanced Manufacturing",
]

# Room assignments, keyed by startup name. Backfill as the mapping comes in.
ROOMS = {
    "Complyt": "MD-1",
    "Sentrix Robotics": "MD-2",
    "GeneGenie": "MD-3",
    "Drivos": "MD-4",
    "VoClyp": "MD-5",
    "Guild": "MD-6",
    "Cloutly": "V-Lab Back",
    "MERU": "GS-3",
    "BRAHMA Aerospace": "GS-2",
    "AI Infra": "V-Lab Front",
    "TRECC": "V-Lab Front",
    "Gridlytics": "V-Lab Center",
    "undark": "V-Lab Center",
}

# Optional website / X links, keyed by startup name. Backfill as they come in.
LINKS = {
    "Guild": {"website": "https://www.guilds.work/", "x": "https://x.com/AdvGuildHQ"},
}

rows = []
for r in range(2, ws.max_row + 1):
    members = cell(r, 1)
    if not members:
        continue
    startup = cell(r, 2)
    rows.append({
        "room": "",
        "startup": startup,
        "members": members,
        "sector": meta[r][0],
        "keywords": meta[r][1],
        "website": "",
        "x": "",
        "problem": cell(r, 3),
        "solution": cell(r, 5),
    })

# Teams not in the Excel sheet, added manually.
EXTRAS = [
    {
        "room": "",
        "startup": "TRECC",
        "members": "Aakash",
        "sector": "Blockchain",
        "keywords": ["DeFi", "Lending", "Agentic payments"],
        "website": "https://trecc.finance",
        "x": "https://x.com/treccfinance",
        "problem": (
            'The "Drunk Chauffeur" Dilemma in AI Finance\n\n'
            "We are rapidly moving into a world where companies and individuals want AI bots to manage money, things like automatically trading crypto, rebalancing investment portfolios, and finding the best yield across decentralized finance markets entirely on autopilot.\n\n"
            "However, there is a massive, high-stakes problem keeping institutional capital on the sidelines:\n\n"
            "AI is Unpredictable: At their core, today's AIs are built on probabilities, not strict logic. They guess the next best action. This means an AI can occasionally \"hallucinate\" or get confused. If a customer service AI hallucinates, it types a funny sentence. If a financial AI hallucinates, it can accidentally send $1,000,000 to the wrong wallet address or buy a worthless asset.\n\n"
            "The \"All-or-Nothing\" Vault Access: To let an AI buy and sell assets, you have to give it the \"keys\" to your digital bank vault. But once the AI has those keys, there is no way to stop it if it goes rogue or makes a mistake. It has the power to drain the entire vault in seconds.\n\n"
            "The Speed vs. Safety Tradeoff: Right now, the only way to prevent this is to have a human review every single transaction the AI wants to make. But having a human constantly click \"Approve\" completely destroys the speed and cost-efficiency of using automated software in the first place."
        ),
        "solution": (
            "The Automated Financial Guardrail\n\n"
            "TRECC acts like a digital, bank-grade security guard that stands between the AI bot and your money.\n\n"
            "Instead of letting the AI talk directly to your digital vault, TRECC sits in the middle. Here is how it protects capital:\n\n"
            "Instant Safety Checks: Every time the AI bot says, \"I want to execute a trade,\" TRECC intercepts that command instantly before any money moves.\n\n"
            "Hard Business Rules: TRECC passes the AI's request through a set of rigid, unbending financial rules set by the business owner (e.g., \"Never risk more than 2% of the portfolio on a single trade,\" or \"Only send money to approved addresses\").\n\n"
            "The Safe Vault: Because the AI never actually gets direct access to the master keys of the vault, it is physically impossible for a confused or rogue AI to drain the funds. It can only execute trades that perfectly match the pre-approved safety limits.\n\n"
            "Sub-Second Execution: This entire safety inspection happens in less than half a second.\n\n"
            "The Bottom Line: TRECC gives businesses and investors the best of both worlds, the lightning-fast execution speed of autonomous AI, with the absolute financial safety and risk management of a traditional bank."
        ),
    },
]
rows.extend(EXTRAS)

# Apply room assignments
for t in rows:
    if t["startup"] in ROOMS:
        t["room"] = ROOMS[t["startup"]]

# Apply link overrides
for t in rows:
    if t["startup"] in LINKS:
        t["website"] = LINKS[t["startup"]].get("website", t["website"])
        t["x"] = LINKS[t["startup"]].get("x", t["x"])

# Stable order: by group_order, preserving sheet order within a group
rows.sort(key=lambda x: group_order.index(x["sector"]))

out = {"groupOrder": group_order, "teams": rows}
with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print("teams:", len(rows))
for g in group_order:
    n = [t["startup"] or t["members"] for t in rows if t["sector"] == g]
    print(f"  {g}: {len(n)} -> {n}")
